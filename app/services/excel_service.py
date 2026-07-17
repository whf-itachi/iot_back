"""Shared Excel generation logic for process log and flatness reports."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series  # type: ignore[unresolved-import]
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

# --- Style constants ---
_THIN_BORDER = Border(
    left=Side(style="thin", color="FF2D3B4F"),
    right=Side(style="thin", color="FF2D3B4F"),
    top=Side(style="thin", color="FF2D3B4F"),
    bottom=Side(style="thin", color="FF2D3B4F"),
)
_FONT_TITLE = Font(name="Microsoft YaHei", bold=True, size=14, color="FF000000")
_FONT_SUBTITLE = Font(name="Microsoft YaHei", size=10, color="FF444444")
_FONT_SECTION = Font(name="Microsoft YaHei", bold=True, size=11, color="FF000000")
_FONT_LABEL = Font(name="Microsoft YaHei", size=10, color="FF000000")
_FONT_VALUE = Font(name="Microsoft YaHei", size=10, bold=True, color="FF000000")
_FONT_UNIT = Font(name="Microsoft YaHei", size=9, color="FF888888")
_FONT_TBL_HDR = Font(name="Microsoft YaHei", bold=True, size=10, color="FF000000")
_FILL_TITLE = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
_FILL_SECTION = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")
_FILL_LABEL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
_FILL_TBL_HDR = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_MILLI_THRESHOLD = 1e12


def _ts_str(ts) -> str:
    """Convert timestamp (ms or s) to readable string."""
    if ts is None:
        return "-"
    n = int(ts)
    if n < _MILLI_THRESHOLD:
        n = n * 1000
    try:
        return datetime.fromtimestamp(n / 1000).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, OSError):
        return str(ts)


def _num(v, d=2):
    """Format numeric value to given decimal places, or return '-' for None."""
    if v is None:
        return "-"
    try:
        return round(float(v), d)
    except (ValueError, TypeError):
        return str(v)


def safe_filename(name: str) -> str:
    """Sanitize string for use as filename."""
    for ch in '<>:"/\\|?*':
        name = name.replace(ch, "_")
    return name.strip() or "unknown"


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def build_process_log_excel(r) -> bytes:
    """Build a single-blade process log Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "加工日志"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14

    row = 1
    # Title
    ws.merge_cells(f"A{row}:C{row}")
    _set_cell(ws, row, 1, "螺栓孔加工结果", _FONT_TITLE, _FILL_TITLE, _ALIGN_CENTER, _THIN_BORDER)
    ws.row_dimensions[row].height = 30
    row += 1

    # Subtitle
    ws.merge_cells(f"A{row}:C{row}")
    sub = f"叶片ID：{r.blade_id or '-'}    设备：{r.device_name or '-'}    上报时间：{_ts_str(r.event_time)}"
    _set_cell(ws, row, 1, sub, _FONT_SUBTITLE, None, _ALIGN_LEFT, _THIN_BORDER)
    row += 1

    def _section(title):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        _set_cell(ws, row, 1, title, _FONT_SECTION, _FILL_SECTION, _ALIGN_LEFT, _THIN_BORDER)
        row += 1

    def _kv(label, value, unit=""):
        nonlocal row
        _set_cell(ws, row, 1, label, _FONT_LABEL, _FILL_LABEL, _ALIGN_RIGHT, _THIN_BORDER)
        _set_cell(ws, row, 2, value, _FONT_VALUE, None, _ALIGN_LEFT, _THIN_BORDER)
        _set_cell(ws, row, 3, unit, _FONT_UNIT, None, _ALIGN_CENTER, _THIN_BORDER)
        row += 1

    _section("基本信息")
    _kv("操作员", r.operator or "-")
    _kv("工厂", r.factory or "-")
    _kv("设备", r.device_type_code or r.device_name or "-")
    _kv("加工开始时间", _ts_str(r.process_start_time))
    _kv("加工结束时间", _ts_str(r.process_end_time))
    _kv("总时长", _num(r.total_duration, 1), "Min")

    _section("扫描结果")
    _kv("扫描结果", r.scan_result or "-")
    _kv("螺栓孔最高点", _num(r.bolt_sleeve_max, 3), "mm")
    _kv("螺栓孔最低点", _num(r.bolt_sleeve_min, 3), "mm")
    _kv("Pitch角度", _num(r.pitch_angle, 3), "°")
    _kv("Yaw角度", _num(r.yaw_angle, 3), "°")
    _kv("BCD预估直径", _num(r.bcd_estimate, 3), "mm")
    _kv("加工前平面度", _num(r.before_flatness, 3), "mm")

    _section("铣磨结果")
    _kv("铣磨深度", _num(r.mill_depth, 1), "mm")
    _kv("铣磨圈数", r.mill_cycles if r.mill_cycles is not None else "-")
    _kv("最终结果", r.mill_result or "-")
    _kv("加工后平面度", _num(r.after_flatness, 3), "mm")

    _section("Process Time")
    _kv("调平和支撑耗时", _num(r.adjust_leg_time, 0), "s")
    _kv("激光调整耗时", _num(r.laser_adjust_time, 0), "s")
    _kv("粗扫耗时", _num(r.rough_scan_time, 0), "s")
    _kv("精扫耗时", _num(r.fine_scan_time, 0), "s")
    _kv("铣磨耗时", _num(r.mill_time, 1), "Min")
    _kv("扫描报告耗时", _num(r.scan_report_time, 0), "s")

    _section("铣磨功率")
    _kv("上部单元平均功率", _num(r.upper_avg_power, 2), "%")
    _kv("上部单元最大功率", _num(r.upper_max_power, 2), "%")
    _kv("下部单元平均功率", _num(r.lower_avg_power, 2), "%")
    _kv("下部单元最大功率", _num(r.lower_max_power, 2), "%")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_flatness_sheet(ws, rec, stage_label: str):
    """Write one flatness sheet (before or after) for a single blade record."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 39
    ws.column_dimensions["C"].width = 39

    hole_angle = rec.hole_angle or []
    hole_value = rec.hole_value or []

    row = 1
    ws.merge_cells(f"A{row}:C{row}")
    _set_cell(ws, row, 1, f"平面度报表（{stage_label}）", _FONT_TITLE, _FILL_TITLE, _ALIGN_CENTER, _THIN_BORDER)
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    sub = f"叶片ID：{rec.blade_id or '-'}    设备：{rec.device_name or '-'}    测量时间：{_ts_str(rec.measure_time)}"
    _set_cell(ws, row, 1, sub, _FONT_SUBTITLE, None, _ALIGN_LEFT, _THIN_BORDER)
    row += 1

    # Statistics section
    ws.merge_cells(f"A{row}:C{row}")
    _set_cell(ws, row, 1, "统计数据", _FONT_SECTION, _FILL_SECTION, _ALIGN_LEFT, _THIN_BORDER)
    row += 1

    for label, val, unit in [
        ("最大值", _num(rec.max_value, 2), "mm"),
        ("最小值", _num(rec.min_value, 2), "mm"),
        ("峰峰值（P-V值）", _num(rec.pv_value, 2), "mm"),
        ("RMS", _num(rec.rms, 2), "mm"),
    ]:
        _set_cell(ws, row, 1, label, _FONT_LABEL, _FILL_LABEL, _ALIGN_RIGHT, _THIN_BORDER)
        _set_cell(ws, row, 2, val, _FONT_VALUE, None, _ALIGN_LEFT, _THIN_BORDER)
        _set_cell(ws, row, 3, unit, _FONT_UNIT, None, _ALIGN_CENTER, _THIN_BORDER)
        row += 1

    # Chart section
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    _set_cell(ws, row, 1, "曲线图", _FONT_SECTION, _FILL_SECTION, _ALIGN_LEFT, _THIN_BORDER)
    row += 1

    chart_reserved_start = row
    row += 16

    # Measurement data table
    if hole_angle:
        ws.merge_cells(f"A{row}:C{row}")
        _set_cell(ws, row, 1, "测量数据", _FONT_SECTION, _FILL_SECTION, _ALIGN_LEFT, _THIN_BORDER)
        row += 1

        for ci, h in enumerate(["#", "孔角度 (°)", "孔测量值 (mm)"], 1):
            _set_cell(ws, row, ci, h, _FONT_TBL_HDR, _FILL_TBL_HDR, _ALIGN_CENTER, _THIN_BORDER)
        row += 1

        data_start_row = row
        for di, angle in enumerate(hole_angle):
            _set_cell(ws, row, 1, di + 1, _FONT_VALUE, None, _ALIGN_CENTER, _THIN_BORDER)
            _set_cell(ws, row, 2, _num(angle, 4), _FONT_VALUE, None, _ALIGN_CENTER, _THIN_BORDER)
            _set_cell(ws, row, 3, _num(hole_value[di] if di < len(hole_value) else None, 4), _FONT_VALUE, None, _ALIGN_CENTER, _THIN_BORDER)
            row += 1
        data_end_row = row - 1

        # Build scatter chart
        chart = ScatterChart()
        chart.width = 19
        chart.height = 8

        chart.graphical_properties = GraphicalProperties()
        chart.graphical_properties.line.noFill = True

        x_vals = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
        y_vals = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_end_row)

        series = Series(y_vals, x_vals)
        series.marker.symbol = "circle"
        series.marker.size = 4
        series.marker.graphicalProperties.solidFill = "60C7F3"
        series.graphicalProperties.line.width = 15000
        series.graphicalProperties.line.solidFill = "60C7F3"
        series.smooth = True
        chart.series.append(series)

        # X-axis (horizontal in Excel)
        chart.x_axis.title = "孔测量值 (mm)"
        chart.x_axis.numFmt = '0.0'
        chart.x_axis.crosses = "min"
        chart.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="E3E6EA"))
        chart.x_axis.majorGridlines = ChartLines(  # type: ignore[assignment]
            spPr=GraphicalProperties(ln=LineProperties(solidFill="F2F3F5"))
        )

        # Y-axis (vertical in Excel)
        chart.y_axis.title = "孔角度 (°)"
        chart.y_axis.numFmt = '0.0'
        chart.y_axis.crosses = "min"
        chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="E3E6EA"))
        chart.y_axis.majorGridlines = ChartLines(  # type: ignore[assignment]
            spPr=GraphicalProperties(ln=LineProperties(solidFill="F2F3F5"))
        )

        chart.legend = None
        ws.add_chart(chart, f"A{chart_reserved_start}")


def build_flatness_excel(before_rec, after_rec) -> bytes:
    """Build a single-blade flatness Excel with before/after in separate sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    if before_rec:
        ws = wb.create_sheet("加工前")
        _build_flatness_sheet(ws, before_rec, "加工前")
    if after_rec:
        ws = wb.create_sheet("加工后")
        _build_flatness_sheet(ws, after_rec, "加工后")

    if not wb.worksheets:
        wb.create_sheet("无数据")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_flatness_single_excel(rec, stage_label: str) -> bytes:
    """Build a flatness Excel with a single sheet (加工前 or 加工后 only)."""
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(stage_label)
    ws = wb[stage_label]

    _build_flatness_sheet(ws, rec, stage_label)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
